
from __future__ import annotations
from io import BytesIO
from copy import copy
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from logic import line_type_from_line_no, compute_summary

TEMPLATE_PATH = Path(__file__).resolve().parent / "Book1.xlsx"
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006", bold=False)
THIN = Side(style="thin", color="000000")

def _copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)

def _copy_row_style(ws_src, ws_dst, src_row: int, dst_row: int, max_col: int):
    if src_row in ws_src.row_dimensions:
        ws_dst.row_dimensions[dst_row].height = ws_src.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        _copy_cell(ws_src.cell(src_row, col), ws_dst.cell(dst_row, col))

def _copy_sheet_dimensions(ws_src, ws_dst):
    for key, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[key].width = dim.width
        ws_dst.column_dimensions[key].hidden = dim.hidden

def _auto_width(ws, min_width=8, max_width=60, extra=2):
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, col).value
            val = "" if val is None else str(val)
            for part in val.split("\n"):
                max_len = max(max_len, len(part))
        ws.column_dimensions[letter].width = min(max(max_len + extra, min_width), max_width)

def _text(v):
    return "" if v is None else str(v)

def build_workbook_bytes(headers_df: pd.DataFrame, raw_lines_df: pd.DataFrame, review_df: pd.DataFrame) -> bytes:
    tpl = load_workbook(TEMPLATE_PATH)
    out = Workbook()
    out.remove(out.active)
    build_asn_sheet(tpl["ASN No"], out.create_sheet("ASN No"), headers_df, raw_lines_df)
    build_header_sheet(tpl["HEADER"], out.create_sheet("HEADER"), headers_df, review_df)
    build_lines_sheet(tpl["LINES"], out.create_sheet("LINES"), headers_df, review_df)
    bio = BytesIO()
    out.save(bio)
    return bio.getvalue()

def build_asn_sheet(ws_tpl, ws_out, headers_df: pd.DataFrame, raw_lines_df: pd.DataFrame):
    _copy_sheet_dimensions(ws_tpl, ws_out)
    for r in range(1, 5):
        _copy_row_style(ws_tpl, ws_out, r, r, ws_tpl.max_column)
    ws_out.merge_cells("A1:K2")
    ws_out["A1"] = "ASN SCANER"
    ws_out["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_out["A1"].font = copy(ws_tpl["A1"].font)

    # Summary by ASN using raw lines
    total_asn = headers_df["ASN No"].nunique() if headers_df is not None and not headers_df.empty else 0
    cpt = op = gp = 0
    if raw_lines_df is not None and not raw_lines_df.empty:
        for asn, g in raw_lines_df.groupby("ASN No", sort=False):
            types = set(g["Line No"].fillna("").map(line_type_from_line_no).tolist())
            if "CPT" in types:
                cpt += 1
            if "OP" in types:
                op += 1
            if "GP" in types:
                gp += 1

    ws_out["N1"] = total_asn
    ws_out["N2"] = cpt
    ws_out["N3"] = op
    ws_out["N4"] = gp
    for cell in ["M1","M2","M3","M4","N1","N2","N3","N4"]:
        ws_out[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_out[cell].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    row_cursor = 5
    if headers_df is None or headers_df.empty:
        _auto_width(ws_out, max_width=24)
        return

    raw_by_asn = {}
    if raw_lines_df is not None and not raw_lines_df.empty:
        for asn, g in raw_lines_df.groupby("ASN No", sort=False):
            raw_by_asn[asn] = g.sort_values("Seq")

    for _, h in headers_df.sort_values(["ASN No"]).iterrows():
        asn = h["ASN No"]
        lines = raw_by_asn.get(asn, pd.DataFrame())
        _copy_row_style(ws_tpl, ws_out, 5, row_cursor, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, 6, row_cursor + 1, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, 7, row_cursor + 2, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, 8, row_cursor + 3, ws_tpl.max_column)

        ws_out.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=11)
        ws_out.cell(row_cursor, 1).value = asn
        ws_out.cell(row_cursor, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_out.cell(row_cursor + 1, 1).value = "DATE"
        ws_out.cell(row_cursor + 1, 2).value = h.get("ETA Date", "")
        ws_out.cell(row_cursor + 2, 1).value = "TIME"
        ws_out.cell(row_cursor + 2, 2).value = h.get("ETA Time", "")

        data_rows = lines.to_dict("records") if not lines.empty else []
        for j, line in enumerate(data_rows):
            r = row_cursor + 4 + j
            _copy_row_style(ws_tpl, ws_out, 9, r, ws_tpl.max_column)
            vals = [
                line.get("Seq"),
                line.get("PO No"),
                line.get("Item No"),
                line.get("Rev"),
                line.get("Quantity"),
                line.get("Uom"),
                line.get("Net Weight (KG)"),
                line.get("Gross Weight (KG)"),
                line.get("Packing Spec."),
                line.get("Lot/MI No./SO No./Invoice No"),
                line.get("Line No"),
            ]
            for c, v in enumerate(vals, start=1):
                ws_out.cell(r, c).value = v
            ws_out.row_dimensions[r].height = 24
        row_cursor += 5 + len(data_rows)

    _auto_width(ws_out, max_width=24)
    for col, width in {"B":14, "C":14, "G":16, "H":16, "J":24, "K":14, "M":12, "N":12}.items():
        ws_out.column_dimensions[col].width = max(ws_out.column_dimensions[col].width or 0, width)

def build_header_sheet(ws_tpl, ws_out, headers_df: pd.DataFrame, review_df: pd.DataFrame):
    _copy_sheet_dimensions(ws_tpl, ws_out)
    _copy_row_style(ws_tpl, ws_out, 1, 1, ws_tpl.max_column)

    first_line_map = {}
    if review_df is not None and not review_df.empty:
        first_line_map = review_df.groupby("ASN No", sort=False)["Line No"].first().to_dict()

    if headers_df is None or headers_df.empty:
        _auto_width(ws_out, max_width=45)
        return

    for idx, (_, h) in enumerate(headers_df.sort_values(["ASN No"]).iterrows(), start=2):
        _copy_row_style(ws_tpl, ws_out, 2, idx, ws_tpl.max_column)
        vals = [
            idx - 1,
            h["ASN No"],
            h["Batch"],
            h["Sold To"],
            h["Bill To"],
            h["Ship To"],
            h["Location"],
            h["ETA"],
            h["ETD"],
            first_line_map.get(h["ASN No"], ""),
        ]
        for c, v in enumerate(vals, start=1):
            ws_out.cell(idx, c).value = v
        for c in range(4, 8):
            ws_out.cell(idx, c).alignment = Alignment(wrap_text=True, vertical="center")
        for c in [1,2,3,8,9,10]:
            ws_out.cell(idx, c).alignment = Alignment(horizontal="center", vertical="center")
        ws_out.row_dimensions[idx].height = 45

    _auto_width(ws_out, max_width=45)
    for col in ['D','E','F','G']:
        ws_out.column_dimensions[col].width = max(ws_out.column_dimensions[col].width or 0, 28)

def build_lines_sheet(ws_tpl, ws_out, headers_df: pd.DataFrame, review_df: pd.DataFrame):
    _copy_sheet_dimensions(ws_tpl, ws_out)
    row_cursor = 1
    if headers_df is None or headers_df.empty:
        _auto_width(ws_out, max_width=32)
        return

    review_df = review_df.copy() if review_df is not None else pd.DataFrame()
    if not review_df.empty and "_Line Type" not in review_df.columns:
        review_df["_Line Type"] = review_df["Line No"].fillna("").map(line_type_from_line_no)

    for idx, (_, h) in enumerate(headers_df.sort_values(["ASN No"]).iterrows()):
        asn = h["ASN No"]
        rows = review_df[review_df["ASN No"] == asn].copy() if not review_df.empty else pd.DataFrame()

        if idx == 0:
            src_title, src_eta, src_blank, src_header, src_data = 1, 2, 3, 4, 5
        else:
            src_title, src_eta, src_blank, src_header, src_data = 9, 10, 11, 12, 13

        _copy_row_style(ws_tpl, ws_out, src_title, row_cursor, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, src_eta, row_cursor + 1, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, src_blank, row_cursor + 2, ws_tpl.max_column)
        _copy_row_style(ws_tpl, ws_out, src_header, row_cursor + 3, ws_tpl.max_column)

        ws_out.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=10)
        ws_out.cell(row_cursor, 1).value = asn
        ws_out.cell(row_cursor, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_out.cell(row_cursor + 1, 1).value = "ETA"
        ws_out.cell(row_cursor + 1, 2).value = h.get("ETA", "")

        type_totals = {"CPT": 0, "OP": 0, "GP": 0}
        if not rows.empty:
            for _, rr in rows.iterrows():
                line_no = str(rr.get("Line No", "") or "").upper().strip()
                total = rr.get("Total Cartons")
                try:
                    total = float(total) if total not in ("", None) else 0
                except Exception:
                    total = 0
                if line_no.startswith("C2"):
                    type_totals["CPT"] += total
                elif line_no.startswith("C1"):
                    type_totals["OP"] += total
                elif line_no.startswith("GP") or "GP JOB" in line_no:
                    type_totals["GP"] += total
        total_cartons = sum(type_totals.values())
        def _pretty_num(v):
            if not v:
                return ""
            return int(v) if abs(v - round(v)) < 1e-9 else round(v, 4)
        ws_out.cell(row_cursor, 13).value = _pretty_num(total_cartons)
        ws_out.cell(row_cursor + 1, 13).value = _pretty_num(type_totals["CPT"])
        ws_out.cell(row_cursor + 2, 13).value = _pretty_num(type_totals["OP"])
        ws_out.cell(row_cursor + 3, 13).value = _pretty_num(type_totals["GP"])
        for rr in [row_cursor, row_cursor+1, row_cursor+2, row_cursor+3]:
            for cc in [12, 13]:
                ws_out.cell(rr, cc).alignment = Alignment(horizontal="center", vertical="center")
                ws_out.cell(rr, cc).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        records = rows.to_dict("records") if not rows.empty else []
        for j, row in enumerate(records, start=1):
            r = row_cursor + 3 + j
            _copy_row_style(ws_tpl, ws_out, src_data, r, ws_tpl.max_column)
            vals = [
                j,
                row.get("Item No"),
                row.get("Rev"),
                row.get("Quantity"),
                row.get("Packing Size"),
                row.get("Full Cartons"),
                row.get("Loose Qty"),
                row.get("Total Cartons"),
                row.get("Line No"),
                row.get("Note"),
            ]
            for c, v in enumerate(vals, start=1):
                ws_out.cell(r, c).value = v
                ws_out.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if _text(row.get("Note")) == "No Packing":
                ws_out.cell(r, 10).fill = RED_FILL
                ws_out.cell(r, 10).font = RED_FONT

        row_cursor += 5 + len(records)

    _auto_width(ws_out, max_width=32)
    for col, width in {"B":14, "D":12, "E":12, "F":12, "G":10, "H":12, "I":18, "J":16, "L":12, "M":12}.items():
        ws_out.column_dimensions[col].width = max(ws_out.column_dimensions[col].width or 0, width)
